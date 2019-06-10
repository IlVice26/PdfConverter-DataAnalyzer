"""
    Attraverso le librerie:
    - pdfConverter.py
    - dataCollect.py
    Salva tutti i dati raccolti all'interno di un file Excel (.xlsx)
    Lavoro svolto per l'azienda Fornace S.R.L
"""

__version__ = '0.0.1'
__author__ = 'Vicentini Elia'

import pdfConverter
import dataCollect
import openpyxl
import time, datetime
import sys


targhe1 = dataCollect.targheEni
targhe2 = dataCollect.targheUnion
targhe3 = dataCollect.targheEsso


def exporttofile():
    
    wb = openpyxl.Workbook()
    ws = wb.active

    cellTarga = ws.cell(row = 1, column = 1)
    cellTarga.value = 'Targa'
    cellDate = ws.cell(row = 1, column = 2)
    cellDate.value = 'Data'
    cellOre = ws.cell(row = 1, column = 3)
    cellOre.value = 'Ora'
    cellCosto = ws.cell(row = 1, column = 4)
    cellCosto.value = 'Pagamento €'
    cellDist = ws.cell(row = 1, column = 5)
    cellDist.value =  "Distributore"

    # Export dei dati di Eni
    keys = list(targhe1.keys())
    for i in range(len(keys)):
        cellRefTarga = ws.cell(row = ws.max_row + 1, column = 1)
        cellRefTarga.value = keys[i]

        for j in range(len(targhe1[keys[i]][0])):
            cellData = ws.cell(row = cellRefTarga.row + j, column = 2)
            cellOra = ws.cell(row = cellRefTarga.row + j, column = 3)
            cellPrezzo = ws.cell(row = cellRefTarga.row + j, column = 4)
            cellDist = ws.cell(row = cellRefTarga.row + j, column = 5)

            cellData.value = targhe1[keys[i]][0][j]
            cellOra.value = targhe1[keys[i]][1][j]
            cellPrezzo.value = targhe1[keys[i]][2][j]
            cellDist.value = targhe1[keys[i]][-1]        

    # Export dei dati di Union
    keys = list(targhe2.keys())
    for i in range(len(keys)):
        refTarga = ws.cell(row = ws.max_row + 1, column = 1)
        refTarga.value = keys[i]

        for j in range(0, len(targhe2[keys[i]][0])):
            # Scrittura dei campi: data, benzina e distributore
            cellData = ws.cell(row = refTarga.row + j, column = 2)
            cellPrezzo = ws.cell(row = refTarga.row + j, column = 4)
            cellDist = ws.cell(row = refTarga.row + j, column = 5)

            cellData.value = targhe2[keys[i]][0][j]
            cellPrezzo.value = targhe2[keys[i]][1][j]
            cellDist.value = targhe2[keys[i]][-1]

    keys = list(targhe3.keys())
    for i in range(len(keys)):
        refTarga = ws.cell(row = ws.max_row + 1, column = 1)
        refTarga.value = keys[i]

        for j in range(0, len(targhe3[keys[i]][0])):
            # Scrittura dei campi: data, benzina e distributore
            cellData = ws.cell(row = refTarga.row + j, column = 2)
            cellPrezzo = ws.cell(row = refTarga.row + j, column = 4)
            cellDist = ws.cell(row = refTarga.row + j, column = 5)

            cellData.value = targhe3[keys[i]][0][j]
            cellPrezzo.value = targhe3[keys[i]][1][j]
            cellDist.value = targhe3[keys[i]][-1]

    ts = time.time()
    s = datetime.datetime.fromtimestamp(ts).strftime(('%d-%m-%Y_%H-%M-%S'))
    exportname = 'output/export_' + s + '.xlsx'

    wb.save(exportname)


def main():
    print('Programma creato da Vicentini Elia per Fornace s.r.l\n')
    print("-- pdfConverter.py -- \n")
    pdfConverter.start()
    print("\n-- dataCollect.py -- \n")
    dataCollect.start()
    print("\n-- exportData.py -- \n")
    print('[>] Esportazione dati in esecuzione')
    exporttofile()
    print('[+] Esportazione dati completata')


if __name__ == "__main__":
    print('Cioa')
