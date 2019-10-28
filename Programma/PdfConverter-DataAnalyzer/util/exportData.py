"""
    Attraverso le librerie:
    - pdfConverter.py
    - dataCollect.py
    Salva tutti i dati raccolti all'interno di un file Excel (.xlsx)
    Lavoro svolto da Elia Vicentini per l'azienda Fornace S.R.L
"""

__version__ = '0.0.1'
__author__ = 'Vicentini Elia'

import util.pdfConverter as pdfConverter
import util.dataCollect as dataCollect
import openpyxl
import time, datetime
import util.setupGui as setupGui
import time


targheDis = dataCollect.targheDistributori
targheLea = dataCollect.targheLeasing
tessere = dataCollect.elemautostrade


def exporttofile():
    """
    Export dati su .xlsx
    """

    # Distributori COGE: 7700170000
    targheDistributori = list(targheDis.keys())
    
    # Leasing COGE: 7700150000
    targheLeasing = list(targheLea.keys())
    
    # Autostrade COGE: 7700409000 
    tessereAutostrade = list(tessere.keys())

    targheDatabase = {}

    # Raccolgo le targhe dal database e associo le info più importanti
    wb = openpyxl.load_workbook(setupGui.PATHWIN32 + '\\database\\database.xlsx')
    wb.active = 0
    ws = wb.active
    for i in range(2, ws.max_row):
        cell = ws.cell(row = i, column = 1)
        if len(str(cell.value)) == 7:
            targa = str(cell.value)
            telepass = ws.cell(row = i, column = 20).value
            viacard = ws.cell(row = i, column = 21).value
            ordine = str(ws.cell(row = i, column = 24).value)

            infotarga = []
            infotarga.append(str(telepass))
            infotarga.append(ordine)
            infotarga.append(str(viacard))

            targheDatabase[targa] = infotarga 
        else:
            pass
    wb.close()

    targheDatabaseKeys = list(targheDatabase.keys())
    
    wb = openpyxl.Workbook()
    wb.active = 0
    ws = wb.active

    # Preparazione file di output
    cdc = ws.cell(row = 1, column = 1).value = 'CDC'
    fornitore = ws.cell(row = 1, column = 2).value = 'FORNITORE'
    coge = ws.cell(row = 1, column = 3).value = 'CO GE'
    valfatt = ws.cell(row = 1, column = 4).value = 'VALORE'
    valuta = ws.cell(row = 1, column = 5).value = 'VALUTA'
    ordine = ws.cell(row = 1, column = 7).value = 'ORDINE'
    targa = ws.cell(row = 1, column = 8).value = 'TARGA'

    for i in range(len(targheDistributori)):
        if targheDistributori[i] in targheDatabaseKeys:
            # Assegnazione dati
            cdc = ws.cell(row = i + 2, column = 1).value = '7300000'
            fornitore = ws.cell(row = i + 2, column = 2).value = '5730000066'
            coge = ws.cell(row = i + 2, column = 3).value = '7700170000'
            valfatt = ws.cell(row = i + 2, column = 4).value = targheDis[targheDistributori[i]]
            valuta = ws.cell(row = i + 2, column = 5).value = 'EUR'
            ordine = ws.cell(row = i + 2, column = 7).value = targheDatabase[targheDistributori[i]][-1]
            targa = ws.cell(row = i + 2, column = 8).value = targheDistributori[i]
        else:
            # Assegnazione dati
            cdc = ws.cell(row = i + 2, column = 1).value = '7300000'
            fornitore = ws.cell(row = i + 2, column = 2).value = '5730000066'
            coge = ws.cell(row = i + 2, column = 3).value = '999999999'
            valfatt = ws.cell(row = i + 2, column = 4).value = targheDis[targheDistributori[i]]
            valuta = ws.cell(row = i + 2, column = 5).value = 'EUR'
            ordine = ws.cell(row = i + 2, column = 7).value = '999999999'
            targa = ws.cell(row = i + 2, column = 8).value = targheDistributori[i]

    tempRow = ws.max_row

    for i in range(0, len(targheLeasing)):
        if targheLeasing[i] in targheDatabaseKeys:
            print('Targa esistente in database: ', targheLeasing[i])

            cdc = ws.cell(row = tempRow + i + 1, column = 1).value = '7300000'
            fornitore = ws.cell(row = tempRow + i + 1, column = 2).value = '5730000066'
            coge = ws.cell(row = tempRow + i + 1, column = 3).value = '7700150000'
            valfatt = ws.cell(row = tempRow + i + 1, column = 4).value = targheLea[targheLeasing[i]]
            valuta = ws.cell(row = tempRow + i + 1, column = 5).value = 'EUR'
            ordine = ws.cell(row = tempRow + i + 1, column = 7).value = targheDatabase[targheLeasing[i]][-1]
            targa = ws.cell(row = tempRow + i + 1, column = 8).value = targheLeasing[i]
        else:
            print('Targa non esistente nel database', targheLeasing[i])
            # Assegnazione dati
            cdc = ws.cell(row = tempRow + i + 1, column = 1).value = '7300000'
            fornitore = ws.cell(row = tempRow + i + 1, column = 2).value = '5730000066'
            coge = ws.cell(row = tempRow + i + 1, column = 3).value = '999999999'
            valfatt = ws.cell(row = tempRow + i + 1, column = 4).value = targheLea[targheLeasing[i]]
            valuta = ws.cell(row = tempRow + i + 1, column = 5).value = 'EUR'
            ordine = ws.cell(row = tempRow + i + 1, column = 7).value = '999999999'
            targa = ws.cell(row = tempRow + i + 1, column = 8).value = targheLeasing[i]

    tempRow = ws.max_row

    for i in range(0, len(tessereAutostrade)):
        targa = ''
        for j in range(len(targheDatabaseKeys)):
            # Controllo corrispondenza apparato telepass
            if str(tessereAutostrade[i]).__contains__(str(targheDatabase[targheDatabaseKeys[j]][0])):
                if not str(targheDatabase[targheDatabaseKeys[j]][0]) == '':
                    targa = targheDatabaseKeys[j]
                    break
            # Se non è un targa, controllo corrispondenza tessera viacard
            elif str(tessereAutostrade[i]).__contains__(str(targheDatabase[targheDatabaseKeys[j]][2])):
                if not str(targheDatabase[targheDatabaseKeys[j]][2]) == '':
                    targa = targheDatabaseKeys[j]
                    break
  
        # Controllo se è un APPARATO TELEPASS o TESSERA VIACARD
        if not tessereAutostrade[i].__contains__('.'):
            if not targa == '':
                cdc = ws.cell(row = tempRow + i + 1, column = 1).value = '7300000'
                fornitore = ws.cell(row = tempRow + i + 1, column = 2).value = '5730000066'
                coge = ws.cell(row = tempRow + i + 1, column = 3).value = '7700409000'
                valfatt = ws.cell(row = tempRow + i + 1, column = 4).value = tessere[tessereAutostrade[i]]
                valuta = ws.cell(row = tempRow + i + 1, column = 5).value = 'EUR'
                ordine = ws.cell(row = tempRow + i + 1, column = 7).value = targheDatabase[targa][-1]
                targa = ws.cell(row = tempRow + i + 1, column = 8).value = targa
            else:
                cdc = ws.cell(row = tempRow + i + 1, column = 1).value = '7300000'
                fornitore = ws.cell(row = tempRow + i + 1, column = 2).value = '5730000066'
                coge = ws.cell(row = tempRow + i + 1, column = 3).value = '7700409000'
                valfatt = ws.cell(row = tempRow + i + 1, column = 4).value = tessere[tessereAutostrade[i]]
                valuta = ws.cell(row = tempRow + i + 1, column = 5).value = 'EUR'
                ordine = ws.cell(row = tempRow + i + 1, column = 7).value = '999999999'
                targa = ws.cell(row = tempRow + i + 1, column = 8).value = targa
        elif tessereAutostrade.__contains__('.') or targa == '':
            cdc = ws.cell(row = tempRow + i + 1, column = 1).value = '7300000'
            fornitore = ws.cell(row = tempRow + i + 1, column = 2).value = '5730000066'
            coge = ws.cell(row = tempRow + i + 1, column = 3).value = '7700409000'
            valfatt = ws.cell(row = tempRow + i + 1, column = 4).value = tessere[tessereAutostrade[i]]
            valuta = ws.cell(row = tempRow + i + 1, column = 5).value = 'EUR'
            ordine = ws.cell(row = tempRow + i + 1, column = 7).value = '999999999'
            targa = ws.cell(row = tempRow + i + 1, column = 8).value = targa

    # Data e ora da aggiungere al file output
    ts = time.time()
    st = datetime.datetime.fromtimestamp(ts).strftime(('%d_%m_%Y_%H_%M_%S'))

    # Export dati su file .xlsx
    try:
        wb.save(setupGui.PATHWIN32 + '\\output\\export_' + str(st) +'.xlsx')
    except PermissionError:
        wb.save(setupGui.PATHWIN32 + '\\output\\export1_' + str(st) + '.xlsx')

def main():
    pdfConverter.start()
    dataCollect.start()
    exporttofile()


if __name__ == "__main__":
    exporttofile()